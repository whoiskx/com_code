from openpyxl import load_workbook, Workbook
from setting import urun


# 读取excel文件
def read_excel(file_name=''):
    # 打开 默认可读写
    wb = load_workbook(file_name)

    # 获取工作表
    # sheet = wb.get_sheet_names()
    sheet = wb.active

    # 获取单元格
    # 获取某个单元格的值，观察excel发现也是先字母再数字的顺序，即先列再行
    # b4 = sheet['B4']
    # cell 3个属性 row, column, coordinate
    for row in sheet.rows:
        ID = row[0].value
        Name = row[1].value
        Layer = row[2].value
        ParentID = row[3].value
        AddOn = row[4].value
        Delete = row[5].value
        timestamp = row[6].value
        data_dict = {
            'ID': ID,
            'Name': Name,
            'Layer': Layer,
            'ParentID': ParentID,
            'AddOn': AddOn,
            'Delete': Delete,
            'timestamp': timestamp,
        }
        urun['Basic'].insert(data_dict)
            # ID = row[0].value
            # ID = row[0].value
            # ID = row[0].value
            # ID = row[0].value
    # 获取行和列
    # sheet.rows    生成器, 每一行的数据，tuple包裹。
    # sheet.columns
    # 获得某行的数据 sheet.rows是生成器类型，不能使用索引，转换成list之后再使用索引
    # list(sheet.rows)[2]
    # 获得任意区间的单元格
    # for i in range(1, 4):
    #     for j in range(1, 3):
    #         print(sheet.cell(row=i, column=j))
    # or
    # for row_cell in sheet['A1':'B3']:
    #     for cell in row_cell:
    #         print(cell)
# 读取excel文件
def read_excel_2(file_name=''):
    # 打开 默认可读写
    wb = load_workbook(file_name)

    # 获取工作表
    sheet = wb.active
    print(sheet.title)
    # 获取单元格
    # 获取某个单元格的值，观察excel发现也是先字母再数字的顺序，即先列再行
    # b4 = sheet['B4']
    # cell 3个属性 row, column, coordinate
    for row in sheet.rows:
        ID = row[0].value
        Name = row[1].value
        Layer = row[2].value
        Category = row[3].value
        AddOn = row[4].value
        timestamp = row[5].value
        GroupID = row[6].value
        data_dict = {
            'ID': ID,
            'Name': Name,
            'Overseas': Layer,
            'Category': Category,
            'AddOn': AddOn,
            'timestamp': timestamp,
            'GroupID': GroupID,
        }
        urun['Basic'].insert(data_dict)

def load_excel():
    # 新建工作表
    wb = Workbook()
    # 获取工作表
    sheet = wb.active
    data = urun.praise_88
    for u in data.find():
        row = [u.get('name'), u.get("praise"), u.get('url')]
        # 写写入单元格  直接赋值：sheet['A1'] = 'good'
        sheet.append(row)
    # 保存文件
    wb.save(r"D:\praise_end1.xlsx")


def main():
    file_name_read = 'to_mongod.xlsx'
    # file_name_read = '22222.xlsx'
    read_excel_2(file_name=file_name_read)
    # load_excel()


if __name__ == '__main__':
    main()
