from openpyxl import load_workbook
from setting import urun
from setting import urun

# wb = load_workbook('example.xls')
wb = load_workbook('facebook20180705.xlsx')
sheet = wb.active

#
# for row in sheet.rows:
#     name, url = '', ''
#     for i, cell in enumerate(row):
#         if i == 3:
#             print(cell.value)
#             name = cell.value
#         if i == 6:
#             print(cell.value)
#             url = cell.value
#     account = {"name": name, 'url': url}
#     urun['test'].insert(account)


# 填写praise
for row in sheet.rows:
    name, url = '', ''
    for i, cell in enumerate(row):
        if i == 3:
            print(cell.value)
            name = cell.value
        if i == 6:
            print(cell.value)
            url = cell.value
    account = {"name": name, 'url': url}
    urun['test'].insert(account)
