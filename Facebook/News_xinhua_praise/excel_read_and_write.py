from openpyxl import load_workbook, Workbook
from setting import urun


# 读取excel文件
def read_excel(file_name=''):
    # 打开 默认可读写
    wb = load_workbook(file_name)

    # 获取工作表
    # sheet = wb.get_sheet_names()
    sheet = wb.active

    # 获取单元格
    # 获取某个单元格的值，观察excel发现也是先字母再数字的顺序，即先列再行
    # b4 = sheet['B4']
    # cell 3个属性 row, column, coordinate
    for row in sheet.rows:
        name, url = '', ''
        for i, cell in enumerate(row):
            if i == 3:
                # print(cell.value)
                # print(cell.row)
                # print(cell.column)
                # print(cell.coordinate)
                name = cell.value
            if i == 6:
                print(cell.value)
                url = cell.value
        account = {
            "name": name,
            'url': url,
        }
        urun.praise_7_11.insert(account)

    # 获取行和列
    # sheet.rows    生成器, 每一行的数据，tuple包裹。
    # sheet.columns
    # 获得某行的数据 sheet.rows是生成器类型，不能使用索引，转换成list之后再使用索引
    # list(sheet.rows)[2]
    # 获得任意区间的单元格
    # for i in range(1, 4):
    #     for j in range(1, 3):
    #         print(sheet.cell(row=i, column=j))
    # or
    # for row_cell in sheet['A1':'B3']:
    #     for cell in row_cell:
    #         print(cell)


def load_excel(file_name=''):
    # 新建工作表
    wb = Workbook()
    # 获取工作表
    sheet = wb.active
    data = urun.praise_7_11_julei_praise_2
    for u in data.find():
        row = [u.get('name'), u.get("praise"), u.get('url')]
        # 写写入单元格  直接赋值：sheet['A1'] = 'good'
        sheet.append(row)
    # 保存文件
    # wb.save(r"D:\praise_7_8_praise_2.xlsx")
    wb.save(file_name)

def main():
    # file_name_read = 'facebook_julei20180711.xlsx'
    # file_name_read = 'facebook20180711.xlsx'
    # read_excel(file_name=file_name_read)

    load_excel(file_name=r'D:\praise_7_11_ju.xlsx')


if __name__ == '__main__':
    main()
