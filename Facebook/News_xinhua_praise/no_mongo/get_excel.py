from openpyxl import load_workbook, Workbook

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

urun = []
result = []

# 读取excel文件
def read_excel(file_name='', db_save='', db_praise=''):
    # 打开 默认可读写
    wb = load_workbook(file_name)

    # 获取工作表
    # sheet = wb.get_sheet_names()
    sheet = wb.active

    # 获取单元格
    # 获取某个单元格的值，观察excel发现也是先字母再数字的顺序，即先列再行
    # b4 = sheet['B4']
    # cell 3个属性 row, column, coordinate
    for row in sheet.rows:
        name, url = '', ''
        for i, cell in enumerate(row):
            if i == 3:
                # print(cell.value)
                # print(cell.row)
                # print(cell.column)
                # print(cell.coordinate)
                name = cell.value
            if i == 6:
                print(cell.value)
                url = cell.value
        account = {
            "name": name,
            'url': url,
        }
        urun.append(account)
    print("get all excel")
    # 获取行和列
    # sheet.rows    生成器, 每一行的数据，tuple包裹。
    # sheet.columns
    # 获得某行的数据 sheet.rows是生成器类型，不能使用索引，转换成list之后再使用索引
    # list(sheet.rows)[2]
    # 获得任意区间的单元格
    # for i in range(1, 4):
    #     for j in range(1, 3):
    #         print(sheet.cell(row=i, column=j))
    # or
    # for row_cell in sheet['A1':'B3']:
    #     for cell in row_cell:
    #         print(cell)
    import time

    from selenium import webdriver

    driver = webdriver.Chrome()

    urls = urun
    error_count = 0
    for count, u in enumerate(urls):
        try:
            if count == 0:
                continue
            url = u.get("url")
            print(url)
            driver.get(url)
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, "_2u_j")))

            praise = driver.find_element_by_class_name('_2u_j').text
            if '次赞' in praise:
                praise = praise.replace('次赞', '')
            print('第{}次'.format(count), praise)
            u.update({'praise': praise})
            result.append(u)
            # if count == 3:
            #     break
        except Exception as e:
            print("===========")
            error_count += 1
            print(error_count, e)
            print(u)
            u.update({'praise': 0})
            result.append(u)
            continue
    print('end')


def load_excel(file_name='', db_praise=''):
    # 新建工作表
    wb = Workbook()
    # 获取工作表
    sheet = wb.active
    for u in result:
        row = [u.get('name'), u.get("praise"), u.get('url')]
        # 写写入单元格  直接赋值：sheet['A1'] = 'good'
        sheet.append(row)
    # 保存文件
    # wb.save(r"D:\praise_7_8_praise_2.xlsx")
    print('save')
    wb.save(file_name)


def main():

    day = '11'
    file_name_read = 'facebook201810{}.xlsx'.format(day)
    db_save = 'save_{}'.format(day)
    db_praise = 'praise_{}'.format(day)
    file_name_load = 'D:\praise_10_{}.xlsx'.format(day)

    read_excel(file_name=file_name_read, db_save=db_save, db_praise=db_praise)
    load_excel(file_name=file_name_load, db_praise=db_praise)


if __name__ == '__main__':
    main()