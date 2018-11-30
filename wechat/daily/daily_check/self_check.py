# -*- coding: utf-8 -*-
import datetime
import json

import requests
from openpyxl import load_workbook, Workbook


class WxCheck(object):
    def __init__(self):
        self.read_name = ''
        self.account_list = []
        self.url = 'http://60.190.238.178:38010/search/common/weixin/select?sort=Time%20desc&Account={}&rows=20&starttime=20180825&endtime=20181130'
        self.write_list = []
        self.save_name = '{}.xlsx'.format(str(datetime.datetime.today().date()))

    def read(self):
        wb = load_workbook(self.read_name)
        sheet = wb.active
        for row in sheet.rows:
            account = row[0].value
            name = row[1].value
            self.account_list.append([account.strip(), name])

    def write(self):
        for account in self.account_list:
            url = self.url.format(account[0])
            resp = requests.get(url)
            result = json.loads(resp.text)
            info = result.get('results')
            if info:
                print(account)
                # 采集时间
                addon = info[0].get('AddOn')
                print(addon)
                # article_time = info[0].get('AddOn')
                addon_date = datetime.datetime.fromtimestamp(int(addon[:-3])).date()
                print(addon_date)
                account.append(str(addon_date))
                # 发文时间
                print(info[0].get('Time'))
                article_time = info[0].get('Time')
                # article_time = info[0].get('AddOn')
                article_date = datetime.datetime.fromtimestamp(int(article_time[:-3])).strftime('%Y-%m-%d %H:%M:%S')
                print(article_date)
                account.append(str(article_date))
                self.write_list.append(account)
            else:
                print("该url底层无数据: ", url)
                account.append("该账号底层无数据")
                self.write_list.append(account)

    def create_wb(self):
        wb = Workbook()
        sheet = wb.active
        sheet.append(['微信ID', '', '最近采集时间', '最近发文时间'])
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 25

        for account in self.write_list:
            sheet.append(account)
        wb.save(self.save_name)

    def run(self):
        self.read()
        self.write()
        self.create_wb()


if __name__ == '__main__':
    test = WxCheck()
    test.read_name = 'account.xlsx'
    test.run()
